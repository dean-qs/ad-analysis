"""Tests for scripts/export.py. Never touch a real OpenAI call --
memo_call is monkeypatched, same pattern as every other module's LLM
boundary.
"""
import json

import pytest

import analyze
import export


def _codebook():
    return {
        "strategy_themes": [
            {"id": "S01", "label": "Access", "definition": "Cost/coverage claims.",
             "inclusion": ["copay"], "exclusion": ["awareness only"],
             "examples": ["Ask about copay support."], "example_ids": ["c1"],
             "brandwatch_boolean_clean": "copay NEAR/5 cost",
             "brandwatch_boolean_labelled": "<<< Access >>>\ncopay NEAR/5 cost"},
        ],
        "topic_themes": [
            {"id": "T01", "label": "Oncology", "definition": "Cancer claims.",
             "inclusion": ["cancer"], "exclusion": [], "examples": [], "example_ids": [],
             "brandwatch_boolean_clean": "cancer", "brandwatch_boolean_labelled": "<<< Oncology >>>\ncancer"},
        ],
    }


def _row(creative_id, brand="Genentech", source="meta", strategy_themes=None,
        topic_themes=None, headline="", body="", transcript="", served_days=5,
        advertiser_id=None, format_="IMAGE"):
    return {"creative_id": creative_id, "brand": brand, "source": source,
            "strategy_themes": strategy_themes or [], "topic_themes": topic_themes or [],
            "headline": headline, "body": body, "transcript": transcript,
            "served_days": served_days, "advertiser_id": advertiser_id, "format": format_}


# ------------------------------------------------------------ limitations


def test_read_limitations_parses_the_real_skill_md():
    """Locks in that the real project file parses, including the
    LinkedIn-dates finding added alongside this module."""
    bullets = export.read_limitations()
    assert any("Meta commercial ads" in b for b in bullets)
    assert any("LinkedIn ad dates" in b for b in bullets)
    assert any("attribution" in b for b in bullets)


def test_read_limitations_missing_file_returns_empty_list(tmp_path):
    assert export.read_limitations(tmp_path / "nope.md") == []


def test_read_limitations_joins_continuation_lines(tmp_path):
    p = tmp_path / "SKILL.md"
    p.write_text("## Known limitations\n\n"
                 "- **Bold thing.** This continues\n"
                 "  onto a second line.\n"
                 "- Second bullet.\n\n## Next section\n")
    bullets = export.read_limitations(p)
    assert bullets == ["**Bold thing.** This continues onto a second line.",
                       "Second bullet."]


# --------------------------------------------------------- ad library url


def test_ad_library_url_meta():
    url = export.ad_library_url(_row("123", source="meta"))
    assert url == "https://www.facebook.com/ads/library/?id=123"


def test_ad_library_url_google_needs_advertiser_id():
    url = export.ad_library_url(_row("CR1", source="google", advertiser_id="AR1"))
    assert url == "https://adstransparency.google.com/advertiser/AR1/creative/CR1"
    assert export.ad_library_url(_row("CR1", source="google", advertiser_id=None)) is None


def test_ad_library_url_linkedin():
    url = export.ad_library_url(_row("999", source="linkedin"))
    assert url == "https://www.linkedin.com/ad-library/detail/999"


def test_ad_library_url_unknown_source_is_none():
    assert export.ad_library_url(_row("1", source="tiktok")) is None


def test_ad_library_url_never_uses_a_relayed_media_url():
    """The whole point: never link to an Apify-relayed, expiring URL.
    media_urls should play no role in this function at all."""
    row = _row("1", source="meta")
    row["media_urls"] = ["https://scontent.fbcdn.net/expiring?oe=6A7A5651"]
    url = export.ad_library_url(row)
    assert "fbcdn" not in url


# ------------------------------------------------------------------ xlsx


def test_build_ads_xlsx_writes_a_readable_file_with_limitations_sheet(tmp_path):
    rows = [_row("c1", strategy_themes=["S01"], topic_themes=["T01"])]
    out_path = tmp_path / "ads.xlsx"
    export.build_ads_xlsx(rows, out_path, ["Limitation one.", "Limitation two."])

    assert out_path.exists()
    import openpyxl
    wb = openpyxl.load_workbook(out_path)
    assert "ads" in wb.sheetnames
    assert "limitations" in wb.sheetnames
    ads_sheet = wb["ads"]
    header = [c.value for c in ads_sheet[1]]
    assert "strategy_themes" in header
    assert "ad_library_url" in header


# -------------------------------------------------------------- codebook.md


def test_codebook_md_includes_both_tiers_booleans_and_limitations():
    md = export.codebook_md(_codebook(), ["A real limitation."])
    assert "S01" in md and "Access" in md
    assert "T01" in md and "Oncology" in md
    assert "<<< Access >>>" in md
    assert "A real limitation." in md
    assert "Strategy themes" in md and "Topic themes" in md


# ------------------------------------------------------------ messaging memo


def test_estimate_memo_counts_real_tokens():
    analysis = {"strategy": {"mix": {"S01": {"Genentech": 0.5}}, "table_stakes": [],
                              "distinctiveness": {}, "register_divergence": {},
                              "platform_register": {}}, "labels": {"S01": "Access"}}
    est = export.estimate_memo(analysis, _codebook())
    assert est["input_tokens"] > 0
    assert est["cost"] > 0


def test_messaging_memo_appends_limitations_verbatim_after_llm_body(monkeypatch):
    def fake_memo_call(system, user):
        usage = type("U", (), {"prompt_tokens": 100, "completion_tokens": 50})()
        return "The headline finding is that Access dominates.", usage
    monkeypatch.setattr(export, "memo_call", fake_memo_call)

    analysis = {"strategy": {"mix": {}, "table_stakes": [], "distinctiveness": {},
                              "register_divergence": {}, "platform_register": {}},
                "labels": {}}
    memo = export.messaging_memo(analysis, _codebook(), ["Verbatim limitation text."])
    assert "The headline finding is that Access dominates." in memo
    assert "Verbatim limitation text." in memo
    assert memo.index("Access dominates") < memo.index("Verbatim limitation text")


def test_messaging_memo_prompt_never_asks_about_meta_spend():
    """The prompt itself must forbid this, not just hope the model
    behaves -- Meta commercial ads genuinely have no spend data."""
    assert "spend" in export.MEMO_SYSTEM_PROMPT.lower()
    assert "does not exist" in export.MEMO_SYSTEM_PROMPT.lower()


def test_messaging_memo_warns_if_model_writes_its_own_limitations(monkeypatch, capsys):
    """Caught live: despite the instruction not to, the model wrote a
    vague self-authored 'Known limitations:' line mid-text, duplicating
    and contradicting the real verbatim section appended after. Can't
    fully prevent this from the prompt alone, so at minimum it must be
    flagged loudly rather than shipped silently."""
    def fake_memo_call(system, user):
        usage = type("U", (), {"prompt_tokens": 10, "completion_tokens": 10})()
        return "Some analysis. Known limitations: this ignores spend data.", usage
    monkeypatch.setattr(export, "memo_call", fake_memo_call)

    analysis = {"strategy": {"mix": {}, "table_stakes": [], "distinctiveness": {},
                              "register_divergence": {}, "platform_register": {}},
                "labels": {}}
    export.messaging_memo(analysis, _codebook(), ["A limitation."])
    assert "WARNING" in capsys.readouterr().out


def test_memo_prompt_does_not_model_dash_heavy_phrasing():
    """Caught live: the prompt's own section list used '--' as a
    separator, and the model dutifully copied that formatting into its
    output section headers -- exactly the dash-heavy style the prompt
    was telling it to avoid. The prompt must not contain what it
    forbids."""
    assert "--" not in export.MEMO_SYSTEM_PROMPT


# --------------------------------------------------------------- explorer


def test_explorer_html_embeds_data_and_has_no_external_network_calls():
    rows = [_row("c1", strategy_themes=["S01"])]
    analysis = {"strategy": {"mix": {"S01": {"Genentech": 1.0}}}, "labels": {"S01": "Access"}}
    html = export.explorer_html(rows, analysis, _codebook(), ["A limitation."])

    assert "__DATA__" not in html   # placeholder got substituted
    assert '"creative_id": "c1"' in html or '"creative_id":"c1"' in html
    assert "A limitation." in html
    assert "<script src=" not in html
    assert 'href="http' not in html   # no external stylesheet/link tags
    assert "cdn." not in html.lower()


def test_explorer_html_theme_filter_includes_both_tiers():
    """The gallery's matches() checks strategy_themes and topic_themes,
    but theme_ids was only ever built from strategy_themes -- topic
    filtering was reachable in code, not in the UI dropdown."""
    rows = [_row("c1", strategy_themes=["S01"], topic_themes=["T01"])]
    analysis = {"strategy": {"mix": {}}, "labels": {"S01": "Access", "T01": "Oncology"}}
    html = export.explorer_html(rows, analysis, _codebook(), [])
    start = html.index("const DATA = ") + len("const DATA = ")
    end = html.index(";\n", start)
    payload = json.loads(html[start:end])
    assert set(payload["theme_ids"]) == {"S01", "T01"}


def test_explorer_html_mix_shape_matches_analyze_pys_real_output(monkeypatch, tmp_path):
    """Caught live: pandas to_json()'s default shape is {brand: {theme:
    value}}, not {theme: {brand: value}}. The explorer's JS originally
    assumed the wrong one and rendered brand names as row labels with
    every cell at 0%. Use analyze.py's real output, not a hand-rolled
    dict, so this can't silently drift again."""
    monkeypatch.chdir(tmp_path)
    rows = [_row("g1", brand="Genentech", strategy_themes=["S01"]),
            _row("p1", brand="Pfizer", strategy_themes=[])]
    result = analyze.analyze_all(rows, _codebook())

    html = export.explorer_html(rows, result, _codebook(), [])
    start = html.index("const DATA = ") + len("const DATA = ")
    end = html.index(";\n", start)
    payload = json.loads(html[start:end])

    assert "Genentech" in payload["mix"]          # brand-keyed at the top
    assert "S01" in payload["mix"]["Genentech"]    # theme-keyed underneath
    assert "renderHeat" in html
    assert "DATA.mix[b] && DATA.mix[b][t]" in html  # locks in the correct access order


def test_explorer_html_is_valid_enough_json_payload():
    rows = [_row("c1", strategy_themes=["S01"], headline="Ask your doctor")]
    analysis = {"strategy": {"mix": {}}, "labels": {}}
    html = export.explorer_html(rows, analysis, _codebook(), [])
    start = html.index("const DATA = ") + len("const DATA = ")
    end = html.index(";\n", start)
    payload = json.loads(html[start:end])
    assert payload["creatives"][0]["creative_id"] == "c1"
    assert payload["creatives"][0]["headline"] == "Ask your doctor"


# -------------------------------------------------------------- build_all


def test_build_all_writes_all_five_paths(monkeypatch, tmp_path):
    def fake_memo_call(system, user):
        usage = type("U", (), {"prompt_tokens": 10, "completion_tokens": 10})()
        return "Body text.", usage
    monkeypatch.setattr(export, "memo_call", fake_memo_call)

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    rows = [_row("c1", strategy_themes=["S01"], topic_themes=["T01"])]
    (out_dir / "ads_coded.json").write_text(json.dumps(rows))
    (out_dir / "codebook.json").write_text(json.dumps(_codebook()))

    paths = export.build_all(out_dir)
    names = {p.name for p in paths}
    assert names == {"ads.xlsx", "codebook.json", "codebook.md",
                     "messaging_memo.md", "explorer.html", "analysis.json"}
    for p in paths:
        assert p.exists()


def test_build_all_writes_analysis_json_into_out_dir_not_cwd(monkeypatch, tmp_path):
    """Caught live: analyze_all() used to hardcode Path("out/analysis.json")
    relative to the current working directory, ignoring build_all's own
    out_dir entirely. It silently leaked a stray file into whatever cwd
    happened to be (the real project directory, every time this test
    suite ran) instead of the out_dir actually being built. Run from a
    cwd that is NOT out_dir's parent, so a regression reappears here
    instead of silently passing again."""
    def fake_memo_call(system, user):
        usage = type("U", (), {"prompt_tokens": 10, "completion_tokens": 10})()
        return "Body text.", usage
    monkeypatch.setattr(export, "memo_call", fake_memo_call)

    unrelated_cwd = tmp_path / "somewhere_else"
    unrelated_cwd.mkdir()
    monkeypatch.chdir(unrelated_cwd)

    out_dir = tmp_path / "real_output"
    out_dir.mkdir()
    rows = [_row("c1", strategy_themes=["S01"])]
    (out_dir / "ads_coded.json").write_text(json.dumps(rows))
    (out_dir / "codebook.json").write_text(json.dumps(_codebook()))

    export.build_all(out_dir)

    assert (out_dir / "analysis.json").exists()
    assert not (unrelated_cwd / "out").exists()


def test_build_all_warns_if_limitations_missing(monkeypatch, tmp_path, capsys):
    def fake_memo_call(system, user):
        usage = type("U", (), {"prompt_tokens": 10, "completion_tokens": 10})()
        return "Body text.", usage
    monkeypatch.setattr(export, "memo_call", fake_memo_call)
    monkeypatch.setattr(export, "read_limitations", lambda *a, **kw: [])

    out_dir = tmp_path / "out"
    out_dir.mkdir()
    (out_dir / "ads_coded.json").write_text(json.dumps([_row("c1")]))
    (out_dir / "codebook.json").write_text(json.dumps(_codebook()))

    export.build_all(out_dir)
    assert "WARNING" in capsys.readouterr().out
